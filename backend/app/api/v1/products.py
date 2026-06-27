"""
Эндпоинты для управления товарами.
"""

from fastapi import APIRouter, HTTPException, Depends, Query, Response, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete
from typing import List, Optional
from pydantic import BaseModel, Field
from loguru import logger
from datetime import datetime, timedelta
import io
import json
import uuid as uuid_pkg

try:
    from openpyxl import Workbook, load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from app.core.database import get_db
from app.core.datetime_utils import utcnow
from app.config import settings
from app.models.product import Product, PriceHistory
from app.models.user import User
from app.api.v1.auth import get_current_user

router = APIRouter()

# Поддерживаемые маркетплейсы для фильтрации и импорта.
ALLOWED_MARKETPLACES = {
    "wildberries",
    "ozon",
    "avito",
    "yandex_market",
    "kazanexpress",
}


def normalize_marketplace_filter(value: Optional[str]) -> Optional[str]:
    """
    Нормализовать и валидировать marketplace из query-параметров.

    Возвращает:
    - None, если фильтр не задан;
    - нормализованное значение, если оно поддерживается.

    Бросает HTTP 400 для неподдерживаемых значений.
    """
    if value is None:
        return None

    normalized_value = str(value).strip().lower()
    if not normalized_value:
        return None

    if normalized_value not in ALLOWED_MARKETPLACES:
        raise HTTPException(
            status_code=400,
            detail=(
                "Некорректный marketplace. Поддерживаются: "
                + ", ".join(sorted(ALLOWED_MARKETPLACES))
            ),
        )

    return normalized_value


class ProductListItemSchema(BaseModel):
    """Схема одной записи товара в списке."""

    id: int
    name: str
    price: float | None = None
    old_price: float | None = None
    category: str | None = None
    brand: str | None = None
    marketplace: str | None = None
    image_url: str | None = None
    is_own: bool
    created_at: str | None = None


class ProductListResponseSchema(BaseModel):
    """Схема ответа для списка товаров с пагинацией."""

    items: List[ProductListItemSchema] = Field(default_factory=list)
    total: int
    limit: int
    offset: int


@router.get(
    "/list",
    response_model=ProductListResponseSchema,
    summary="Список товаров с фильтрами",
    description="Получение списка товаров с фильтрацией, поиском и сортировкой. Возвращает только товары текущего пользователя.",
)
async def get_products_list(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    search: Optional[str] = Query(None, description="Поиск по названию"),
    marketplace: Optional[str] = Query(None, description="Фильтр по маркетплейсу"),
    is_own: Optional[bool] = Query(None, description="Свои товары (True) или конкуренты (False)"),
    sort_by: str = Query("created_at", pattern="^(created_at|name|price)$", description="Поле для сортировки"),
    sort_order: str = Query("desc", pattern="^(asc|desc)$", description="Порядок: asc или desc"),
    limit: int = Query(100, ge=1, le=500, description="Лимит записей"),
    offset: int = Query(0, ge=0, description="Смещение"),
):
    """
    Получение списка товаров с фильтрами.
    
    Поддерживает:
    - Поиск по названию
    - Фильтр по маркетплейсу
    - Фильтр: свои/конкуренты
    - Сортировка по разным полям
    """
    normalized_marketplace = normalize_marketplace_filter(marketplace)

    logger.info(
        f"📦 Запрос товаров: search={search}, marketplace={normalized_marketplace}, user_id={current_user.id}"
    )

    # Собираем единый набор фильтров, чтобы одинаково применять их и к count, и к выборке.
    filters = [Product.user_id == current_user.id]

    # Применяем фильтры.
    if search:
        filters.append(Product.name.ilike(f"%{search}%"))

    if normalized_marketplace:
        filters.append(Product.marketplace == normalized_marketplace)

    if is_own is not None:
        filters.append(Product.is_competitor == (not is_own))

    # Отдельно считаем общее количество записей по фильтрам (для корректной пагинации на фронтенде).
    total_query = select(func.count(Product.id)).where(*filters)
    total_result = await db.execute(total_query)
    total_count = total_result.scalar() or 0

    query = select(Product).where(*filters)

    # Сортировка только по разрешённым полям.
    allowed_sort_fields = {
        "created_at": Product.created_at,
        "name": Product.name,
        "price": Product.price,
    }
    normalized_sort_by = (sort_by or "created_at").strip().lower()
    normalized_sort_order = (sort_order or "desc").strip().lower()

    order_column = allowed_sort_fields.get(normalized_sort_by, Product.created_at)
    if normalized_sort_order == "desc":
        query = query.order_by(order_column.desc())
    else:
        query = query.order_by(order_column.asc())

    # Лимит и смещение.
    query = query.limit(limit).offset(offset)

    # Выполняем запрос.
    result = await db.execute(query)
    products = result.scalars().all()
    
    # Конвертируем в dict
    items = [
        {
            "id": p.id,
            "name": p.name,
            "price": p.price,
            "old_price": p.old_price,
            "category": p.category,
            "brand": p.brand,
            "marketplace": p.marketplace,
            "image_url": p.images[0] if p.images and isinstance(p.images, list) and len(p.images) > 0 else None,
            "is_own": not p.is_competitor,
            "created_at": p.created_at.isoformat() if p.created_at else None,
        }
        for p in products
    ]

    return {
        "items": items,
        "total": total_count,
        "limit": limit,
        "offset": offset,
    }


@router.get(
    "/{product_id:int}",
    response_model=dict,
    summary="Товар по ID",
    description="Получение детальной информации о товаре. Только владелец может просмотреть свой товар.",
)
async def get_product(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Получение товара по ID.
    """
    logger.info(f"📦 Запрос товара ID={product_id}, user_id={current_user.id}")
    
    result = await db.execute(select(Product).where(Product.id == product_id, Product.user_id == current_user.id))
    product = result.scalar_one_or_none()
    
    if not product:
        raise HTTPException(status_code=404, detail="Товар не найден")
    
    return {
        "id": product.id,
        "external_id": product.external_id,
        "marketplace": product.marketplace,
        "name": product.name,
        "description": product.description,
        "price": product.price,
        "old_price": product.old_price,
        "discount": product.discount,
        "rating": product.rating,
        "reviews_count": product.reviews_count,
        "sales_count": product.sales_count,
        "category": product.category,
        "brand": product.brand,
        "images": product.images,
        "characteristics": product.characteristics,
        "is_competitor": product.is_competitor,
        "url": product.url,
        "last_parsed_at": product.last_parsed_at.isoformat() if product.last_parsed_at else None,
    }


@router.get(
    "/stats/summary",
    response_model=dict,
    summary="Статистика по товарам",
    description="Общая статистика за период: количество товаров, средняя цена, дельта к предыдущему периоду и т.д. Только для текущего пользователя.",
)
async def get_product_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    period: str = Query("30d", description="Период: today, 7d, 30d"),
):
    """
    Получение статистики по товарам с учётом выбранного периода.

    Логика:
    - Текущий период: [start_at, now]
    - Предыдущий период: [start_at - duration, start_at)
    """
    logger.info(f"📊 Запрос статистики товаров: user_id={current_user.id}, period={period}")

    # Нормализуем период и ограничиваем список поддерживаемых значений.
    normalized_period = (period or "30d").strip().lower()
    period_to_delta = {
        "today": timedelta(days=1),
        "7d": timedelta(days=7),
        "30d": timedelta(days=30),
    }

    if normalized_period not in period_to_delta:
        raise HTTPException(
            status_code=400,
            detail="Некорректный period. Поддерживаются: today, 7d, 30d",
        )

    period_delta = period_to_delta[normalized_period]
    now = utcnow()
    current_period_start = now - period_delta
    previous_period_start = current_period_start - period_delta

    # Базовый фильтр по владельцу.
    user_filter = Product.user_id == current_user.id
    
    # Фильтр для текущего периода.
    current_period_filter = (
        Product.created_at >= current_period_start,
        Product.created_at <= now,
    )

    # Фильтр для предыдущего периода.
    previous_period_filter = (
        Product.created_at >= previous_period_start,
        Product.created_at < current_period_start,
    )
    
    # ========== Текущий период ==========
    total_query = select(func.count(Product.id)).where(user_filter, *current_period_filter)
    total_result = await db.execute(total_query)
    total_count = total_result.scalar() or 0

    competitor_query = select(func.count(Product.id)).where(
        user_filter,
        *current_period_filter,
        Product.is_competitor == True,
    )
    competitor_result = await db.execute(competitor_query)
    competitor_count = competitor_result.scalar() or 0

    own_count = max(0, total_count - competitor_count)

    avg_price_query = select(func.avg(Product.price)).where(user_filter, *current_period_filter)
    avg_price_result = await db.execute(avg_price_query)
    avg_price = avg_price_result.scalar() or 0

    avg_competitor_price_query = select(func.avg(Product.price)).where(
        user_filter,
        *current_period_filter,
        Product.is_competitor == True,
    )
    avg_competitor_price_result = await db.execute(avg_competitor_price_query)
    avg_competitor_price = avg_competitor_price_result.scalar() or 0

    marketplace_query = select(
        Product.marketplace,
        func.count(Product.id),
    ).where(user_filter, *current_period_filter).group_by(Product.marketplace)
    marketplace_result = await db.execute(marketplace_query)
    marketplace_dist = {mp: count for mp, count in marketplace_result.all()}

    discount_query = select(func.count(Product.id)).where(
        user_filter,
        *current_period_filter,
        Product.discount > 0,
    )
    discount_result = await db.execute(discount_query)
    discount_count = discount_result.scalar() or 0

    # ========== Предыдущий период (для delta) ==========
    previous_total_query = select(func.count(Product.id)).where(user_filter, *previous_period_filter)
    previous_total_result = await db.execute(previous_total_query)
    previous_total_count = previous_total_result.scalar() or 0

    previous_avg_price_query = select(func.avg(Product.price)).where(user_filter, *previous_period_filter)
    previous_avg_price_result = await db.execute(previous_avg_price_query)
    previous_avg_price = previous_avg_price_result.scalar() or 0

    def calculate_delta_percent(current_value: float, previous_value: float) -> float | None:
        """
        Рассчитать изменение в процентах относительно предыдущего периода.

        Возвращает None, если нет корректной базы для сравнения.
        """
        if previous_value is None or previous_value <= 0:
            return None
        return round(((current_value - previous_value) / previous_value) * 100, 2)

    return {
        "period": normalized_period,
        "period_start": current_period_start.isoformat(),
        "period_end": now.isoformat(),
        "total_products": total_count,
        "competitor_products": competitor_count,
        "own_products": own_count,
        "average_price": round(avg_price, 2) if avg_price else 0,
        "average_competitor_price": round(avg_competitor_price, 2) if avg_competitor_price else 0,
        "marketplace_distribution": marketplace_dist,
        "products_with_discount": discount_count,
        "previous_period": {
            "total_products": previous_total_count,
            "average_price": round(previous_avg_price, 2) if previous_avg_price else 0,
        },
        "delta": {
            "total_products_percent": calculate_delta_percent(float(total_count), float(previous_total_count)),
            "average_price_percent": calculate_delta_percent(float(avg_price or 0), float(previous_avg_price or 0)),
        },
    }


@router.delete(
    "/{product_id:int}",
    status_code=204,
    summary="Удаление товара",
    description="Удаление товара по ID. Только владелец может удалить свой товар.",
)
async def delete_product(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Удаление товара.
    """
    logger.info(f"🗑️ Удаление товара ID={product_id}, user_id={current_user.id}")
    
    # Проверяем товар сразу с учётом владельца, чтобы не раскрывать наличие чужих ID.
    result = await db.execute(
        select(Product).where(
            Product.id == product_id,
            Product.user_id == current_user.id,
        )
    )
    product = result.scalar_one_or_none()
    
    if not product:
        raise HTTPException(status_code=404, detail="Товар не найден")
    
    await db.execute(
        delete(Product).where(
            Product.id == product_id,
            Product.user_id == current_user.id,
        )
    )
    await db.commit()
    
    return None


@router.get(
    "/export",
    summary="Экспорт товаров в Excel",
    description="Выгрузка списка товаров в XLSX формат. Только товары текущего пользователя.",
)
async def export_products(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    marketplace: Optional[str] = Query(None, description="Фильтр по маркетплейсу"),
    is_own: Optional[bool] = Query(None, description="Свои товары или конкуренты"),
):
    """
    Экспорт товаров в Excel файл.
    
    Генерирует XLSX файл со списком товаров.
    """
    normalized_marketplace = normalize_marketplace_filter(marketplace)

    logger.info(
        f"📊 Экспорт товаров в Excel: user_id={current_user.id}, marketplace={normalized_marketplace}"
    )

    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Библиотека openpyxl не установлена")
    
    query = select(Product).where(Product.user_id == current_user.id)
    
    # Фильтры
    if normalized_marketplace:
        query = query.where(Product.marketplace == normalized_marketplace)
    
    if is_own is not None:
        query = query.where(Product.is_competitor == (not is_own))
    
    # Сортировка по дате
    query = query.order_by(Product.created_at.desc())
    
    result = await db.execute(query)
    products = result.scalars().all()
    
    # Создаём Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    
    # Заголовки
    headers = [
        "ID", "Название", "Бренд", "Категория", "Цена", "Старая цена",
        "Скидка %", "Рейтинг", "Отзывы", "Маркетплейс", "Тип", "Дата добавления"
    ]
    ws.append(headers)

    # Данные
    for p in products:
        ws.append([
            p.id,
            p.name,
            p.brand or "",
            p.category or "",
            p.price or 0,
            p.old_price or "",
            p.discount or "",
            p.rating or "",
            p.reviews_count or 0,
            p.marketplace or "",
            "Свои" if not p.is_competitor else "Конкуренты",
            p.created_at.strftime("%d.%m.%Y %H:%M") if p.created_at else ""
        ])
    
    # Форматирование
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, ValueError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в буфер
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"megashark_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get(
    "/import/template",
    summary="Скачать шаблон Excel для импорта",
    description="Возвращает XLSX шаблон с корректными заголовками и примерами строк для импорта товаров.",
)
async def download_import_template():
    """
    Скачать готовый XLSX-шаблон для импорта товаров.

    Шаблон содержит:
    - обязательные колонки;
    - опциональные колонки;
    - несколько примеров заполнения.
    """
    logger.info("📄 Запрос шаблона импорта")

    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Библиотека openpyxl не установлена")

    wb = Workbook()
    ws = wb.active
    ws.title = "Import Template"

    headers = ["name", "price", "marketplace", "category", "brand", "is_own"]
    ws.append(headers)

    sample_rows = [
        ["Футболка базовая", 1999, "wildberries", "Одежда", "MegaBrand", "true"],
        ["Кроссовки повседневные", 5490, "ozon", "Обувь", "RunFast", "false"],
        ["Рюкзак городской", 3290, "yandex_market", "Аксессуары", "CityPack", "1"],
    ]

    for row in sample_rows:
        ws.append(row)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            cell_value = "" if cell.value is None else str(cell.value)
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = "megashark_import_template.xlsx"

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post(
    "/import",
    summary="Импорт товаров из Excel",
    description="Массовая загрузка товаров из XLSX файла. Товары закрепляются за текущим пользователем.",
)
async def import_products(
    file: UploadFile = File(..., description="XLSX файл с товарами"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Импорт товаров из Excel файла.
    
    Ожидаемые колонки:
    - name (обязательно)
    - price (обязательно)
    - marketplace (wildberries, ozon, avito, yandex_market, kazanexpress)
    - category
    - brand
    - is_own (true/false/1/0/yes/no)
    """
    logger.info(f"📥 Импорт товаров из Excel: user_id={current_user.id}")
    
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Библиотека openpyxl не установлена")
    
    file_name = (file.filename or "").lower().strip()
    if not file_name.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Поддерживается только формат .xlsx")

    def parse_bool_cell(raw_value: object, default_value: bool = False) -> bool:
        """Преобразовать значение ячейки в bool с поддержкой строковых форматов."""
        if raw_value is None:
            return default_value

        if isinstance(raw_value, bool):
            return raw_value

        if isinstance(raw_value, (int, float)):
            return raw_value != 0

        normalized_value = str(raw_value).strip().lower()
        if not normalized_value:
            return default_value

        if normalized_value in {"1", "true", "yes", "y", "да"}:
            return True
        if normalized_value in {"0", "false", "no", "n", "нет"}:
            return False

        return default_value

    try:
        content = await file.read()
        if len(content) > settings.max_import_file_bytes:
            raise HTTPException(
                status_code=400,
                detail=f"Файл слишком большой (макс. {settings.max_import_file_bytes // (1024 * 1024)} МБ)",
            )
        if not content:
            raise HTTPException(status_code=400, detail="Файл пустой")

        buffer = io.BytesIO(content)
        wb = load_workbook(buffer, data_only=True)
        ws = wb.active
        
        if ws.max_row < 1:
            raise HTTPException(status_code=400, detail="Файл не содержит заголовков")

        headers = [cell.value for cell in ws[1]]
        header_map = {
            str(header).strip().lower(): index
            for index, header in enumerate(headers)
            if header is not None and str(header).strip()
        }
        
        required = ["name", "price"]
        for field in required:
            if field not in header_map:
                raise HTTPException(
                    status_code=400,
                    detail=f"Отсутствует обязательная колонка: {field}",
                )
        
        products_to_add: list[Product] = []
        row_errors: list[dict] = []
        total_rows = 0
        skipped_rows = 0
        max_reported_errors = 30

        def get_cell_value(row_data: tuple, column_name: str, default_value: object = None) -> object:
            """Безопасно получить значение ячейки по имени колонки."""
            column_index = header_map.get(column_name)
            if column_index is None:
                return default_value
            if column_index >= len(row_data):
                return default_value
            return row_data[column_index]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            total_rows += 1

            # Полностью пустые строки пропускаем без добавления в ошибки.
            if row is None or all((cell is None or str(cell).strip() == "") for cell in row):
                skipped_rows += 1
                continue
        
            current_row_errors: list[str] = []

            raw_name = get_cell_value(row, "name")
            normalized_name = str(raw_name).strip() if raw_name is not None else ""
            if not normalized_name:
                current_row_errors.append("Пустое поле name")

            raw_price = get_cell_value(row, "price")
            normalized_price: float | None = None
            if raw_price is None or str(raw_price).strip() == "":
                current_row_errors.append("Пустое поле price")
            else:
                try:
                    normalized_price = float(raw_price)
                    if normalized_price <= 0:
                        current_row_errors.append("Цена должна быть больше 0")
                except (ValueError, TypeError):
                    current_row_errors.append("Некорректное значение price")

            raw_marketplace = get_cell_value(row, "marketplace", "wildberries")
            normalized_marketplace = "wildberries"
            try:
                normalized_marketplace = normalize_marketplace_filter(str(raw_marketplace or "wildberries")) or "wildberries"
            except HTTPException:
                current_row_errors.append(
                    "Некорректный marketplace. Поддерживаются: " + ", ".join(sorted(ALLOWED_MARKETPLACES))
                )

            raw_category = get_cell_value(row, "category")
            normalized_category = str(raw_category).strip() if raw_category is not None else None
            if normalized_category == "":
                normalized_category = None

            raw_brand = get_cell_value(row, "brand")
            normalized_brand = str(raw_brand).strip() if raw_brand is not None else None
            if normalized_brand == "":
                normalized_brand = None

            raw_is_own = get_cell_value(row, "is_own", False)
            normalized_is_own = parse_bool_cell(raw_is_own, default_value=False)

            if current_row_errors:
                skipped_rows += 1
                logger.warning(f"❌ Ошибка в строке {row_idx}: {'; '.join(current_row_errors)}")

                if len(row_errors) < max_reported_errors:
                    row_errors.append(
                        {
                            "row": row_idx,
                            "errors": current_row_errors,
                        }
                    )
                continue

            products_to_add.append(
                Product(
                    name=normalized_name,
                    price=normalized_price,
                    marketplace=normalized_marketplace,
                    category=normalized_category,
                    brand=normalized_brand,
                    is_competitor=not normalized_is_own,
                    user_id=current_user.id,
                )
            )

        if products_to_add:
            db.add_all(products_to_add)
            await db.commit()
        
        imported_count = len(products_to_add)

        if total_rows == 0:
            raise HTTPException(
                status_code=400,
                detail="В файле нет строк для импорта (кроме заголовков)",
            )

        if imported_count == 0 and total_rows > 0:
            raise HTTPException(
                status_code=400,
                detail="Не удалось импортировать ни одной строки. Проверьте формат файла.",
            )

        response_status = "success" if skipped_rows == 0 else "partial_success"
        response_message = (
            f"Импортировано {imported_count} из {total_rows} строк"
            if skipped_rows > 0
            else f"Импортировано {imported_count} товаров"
        )

        return {
            "status": response_status,
            "imported": imported_count,
            "skipped": skipped_rows,
            "total_rows": total_rows,
            "errors": row_errors,
            "message": response_message,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Ошибка импорта: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка импорта: {str(e)}")


@router.get(
    "/{product_id:int}/price-history",
    response_model=List[dict],
    summary="История цен товара",
    description="Получение истории изменения цен товара",
)
async def get_price_history(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    limit: int = Query(30, ge=1, le=100, description="Количество записей")
):
    """
    Получение истории цен товара.
    
    Возвращает последние N записей об изменении цены.
    """
    logger.info(f"📈 Запрос истории цен для товара ID={product_id}, user_id={current_user.id}")

    # Проверяем, что товар принадлежит текущему пользователю.
    product_result = await db.execute(
        select(Product.id).where(
            Product.id == product_id,
            Product.user_id == current_user.id,
        )
    )
    product_exists = product_result.scalar_one_or_none()
    
    if not product_exists:
        raise HTTPException(status_code=404, detail="Товар не найден")
    
    query = (
        select(PriceHistory)
        .where(PriceHistory.product_id == product_id)
        .order_by(PriceHistory.created_at.desc())
        .limit(limit)
    )
    
    result = await db.execute(query)
    history = result.scalars().all()
    
    return [
        {
            "id": h.id,
            "price": h.price,
            "old_price": h.old_price,
            "discount": h.discount,
            "created_at": h.created_at.isoformat() if h.created_at else None,
        }
        for h in history
    ]
